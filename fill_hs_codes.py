"""
Vult automatisch HS-codes in een Excel-invoice op basis van artikelnummer.

Gebruik CLI:
    python fill_hs_codes.py invoice.xlsx output.xlsx

De mapping staat standaard in hs_mapping.csv naast dit script.
"""
from __future__ import annotations

import csv
import re
import runpy
import sys
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

ARTICLE_HEADERS = {
    "article no",
    "article no.",
    "art no",
    "art no.",
    "artikelnummer",
    "article number",
    "model",
    "buyer art no",
    "buyer art no.",
    "buyer article no",
    "buyer article no.",
    "item no",
    "item no.",
    "qhp item no",
    "qhp item no.",
    "stockcode item qhp",
    "stockcode qhp",
    "our product code party's code",
    "our product code partys code",
    "party's code",
    "partys code",
}
HS_HEADERS = {"hs code (artikellijst)", "hscode (artikellijst)", "hs-code (artikellijst)"}
SIZE_HEADERS = {"size", "maat"}


def create_rapidocr_engine():
    rapidocr_error = None
    try:
        from rapidocr import RapidOCR
    except Exception as exc:
        rapidocr_error = exc
        try:
            from rapidocr_onnxruntime import RapidOCR
        except Exception as fallback_exc:
            raise ImportError(f"rapidocr: {rapidocr_error}; rapidocr_onnxruntime: {fallback_exc}") from fallback_exc

    return RapidOCR()


def iter_rapidocr_results(ocr, image):
    result = ocr(image)
    if isinstance(result, tuple):
        for item in result[0] or []:
            yield item
        return

    boxes = getattr(result, "boxes", None)
    texts = getattr(result, "txts", None)
    scores = getattr(result, "scores", None)
    if boxes is None or texts is None or scores is None:
        return

    for box, text, score in zip(boxes, texts, scores):
        yield box, text, score


def norm_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_key(value) -> str:
    text = norm_text(value).upper()
    text = text.replace("\u00a0", " ")
    text = text.replace(".", "")
    text = re.sub(r"\s+", " ", text)
    return text


def compact_key(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", norm_key(value))


def norm_header(value) -> str:
    return re.sub(r"\s+", " ", norm_text(value).lower().replace(":", "").replace(".", ""))


def norm_hs(value) -> str:
    text = norm_text(value)
    if not text:
        return ""
    # Excel may store HS codes as floats/scientific notation in source files.
    try:
        if re.fullmatch(r"[0-9]+(\.0+)?", text):
            text = str(int(float(text)))
    except Exception:
        pass
    digits = re.sub(r"\D", "", text)
    return digits


def norm_size(value) -> str:
    text = norm_key(value)
    text = re.sub(r"\b(CM|MTR|METER)\b", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_catalog(mapping_csv: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    mapping: Dict[str, str] = {}
    names: Dict[str, str] = {}

    def add_mapping_alias(alias: str, hs: str, name: str = "") -> None:
        alias = norm_key(alias)
        if not alias or not hs:
            return
        mapping.setdefault(alias, hs)
        if name:
            names.setdefault(alias, name)
        compact_alias = compact_key(alias)
        if compact_alias:
            mapping.setdefault(compact_alias, hs)
            if name:
                names.setdefault(compact_alias, name)

    with mapping_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            article = norm_key(row.get("artikelnummer") or row.get("article") or row.get("article_no"))
            name = norm_text(row.get("omschrijving") or row.get("description") or row.get("article_name"))
            hs = norm_hs(row.get("hs_code") or row.get("HS code") or row.get("hs"))
            if article and hs:
                add_mapping_alias(article, hs, name)

                variant = re.sub(r"\s+\d{1,3}(?:-\d{1,3})?$", "", article).strip()
                if variant and variant != article and len(variant.split()) >= 2:
                    variant_name = re.sub(r"\s+\d{1,3}(?:-\d{1,3})?$", "", name).strip()
                    add_mapping_alias(variant, hs, variant_name)
    return mapping, names


def load_mapping(mapping_csv: Path) -> Dict[str, str]:
    mapping, _names = load_catalog(mapping_csv)
    return mapping


def extract_pdf_quantity_amount(text: str) -> Tuple[str, str, str]:
    # Matrix invoices can contain many size/order numbers on the left. The invoice
    # totals requested by the user live in the right-hand quantity/amount columns.
    right_side = (text or "")[180:]
    numbers = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", right_side)
    if len(numbers) < 3:
        return "", "", ""

    qty, _unit_price, amount = numbers[-3:]
    return qty.replace(",", ""), _unit_price, amount


def extract_pdf_quantity_amount_for_article(text: str, article: str) -> Tuple[str, str, str]:
    raw_text = text or ""
    article_regex = re.escape(norm_text(article)).replace(r"\ ", r"\s+")
    match = re.search(rf"^\s*{article_regex}\s+(?P<quantity>\d+)\b", raw_text, flags=re.I)
    numbers = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", raw_text)
    if match and len(numbers) >= 2:
        return match.group("quantity"), numbers[-2], numbers[-1]
    return extract_pdf_quantity_amount(raw_text)


def enrich_pdf_rows(rows: List[dict], names: Dict[str, str]) -> List[dict]:
    for row in rows:
        article = row.get("article", "")
        lookup_article = row.get("lookup_article", article)
        quantity = row.get("quantity", "")
        unit_price = row.get("unit_price", "")
        amount = row.get("amount", "")
        if not (quantity and unit_price and amount) and not row.get("skip_value_fallback"):
            quantity, unit_price, amount = extract_pdf_quantity_amount_for_article(
                row.get("raw_text") or row.get("text", ""),
                article,
            )
        row["article_name"] = (
            names.get(article)
            or names.get(compact_key(article))
            or names.get(lookup_article)
            or names.get(compact_key(lookup_article))
            or ""
        )
        row["quantity"] = quantity
        row["unit_price"] = unit_price
        row["amount"] = amount

    return rows


def article_group_key(article: str) -> str:
    match = re.search(r"\b\d{3,5}\b", norm_key(article))
    return match.group(0) if match else norm_key(article)


def parse_pdf_number(value: str):
    text = re.sub(r"[^0-9,.\-]", "", norm_text(value))
    if not text:
        return ""
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) <= 3:
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def article_pattern(article: str) -> re.Pattern:
    escaped = re.escape(article).replace(r"\ ", r"\s+")
    return re.compile(rf"(?<![A-Z0-9]){escaped}(?![A-Z0-9])")


def norm_pdf_token(value: str) -> str:
    return norm_text(value).lower().replace(":", "").rstrip(".")


def get_pdf_page_text(page) -> str:
    try:
        return page.extract_text(extraction_mode="layout") or ""
    except TypeError:
        return page.extract_text() or ""


def is_packing_list_page(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", text or "").strip().lower()
    compact = re.sub(r"[^a-z]", "", normalized)
    return "packing list" in normalized or "packlist" in compact or "装箱单" in (text or "")


def find_headers(ws) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Return row, article_col, hs_col, size_col. 1-based indexes."""
    best = (None, None, None, None)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80)):
        article_col = None
        hs_col = None
        size_col = None
        for cell in row:
            h = norm_header(cell.value)
            if h in ARTICLE_HEADERS:
                article_col = cell.column
            if h in HS_HEADERS:
                hs_col = cell.column
            if h in SIZE_HEADERS:
                size_col = cell.column
        if article_col:
            best = (row[0].row, article_col, hs_col, size_col)
            break
    return best


def lookup_hs(mapping: Dict[str, str], article_value, size_value=None) -> Optional[str]:
    article = norm_key(article_value)
    if not article:
        return None

    candidates = [article]
    size = norm_size(size_value)
    if size:
        candidates.insert(0, norm_key(f"{article} {size}"))

    for candidate in candidates:
        hs = mapping.get(candidate) or mapping.get(compact_key(candidate))
        if hs:
            return hs
    return None


def lookup_article_from_excel_cell(mapping: Dict[str, str], article_value, size_value=None) -> Tuple[Optional[str], Optional[str]]:
    article = norm_key(article_value)
    if not article:
        return None, None

    size = norm_size(size_value)
    candidates = [article]
    if size:
        candidates.insert(0, norm_key(f"{article} {size}"))

    for candidate in candidates:
        direct = mapping.get(candidate) or mapping.get(compact_key(candidate))
        if direct:
            return candidate, direct

    text_article = find_article_in_text(article, mapping, []) or find_article_prefix(article, mapping)
    if text_article:
        return text_article, mapping.get(text_article) or mapping.get(compact_key(text_article))

    compact = compact_key(article)
    for mapped_article in sorted(mapping, key=lambda item: len(compact_key(item)), reverse=True):
        mapped_compact = compact_key(mapped_article)
        if len(mapped_compact) < 4 or mapped_compact.isdigit():
            continue
        if mapped_compact in compact:
            return mapped_article, mapping.get(mapped_article) or mapping.get(mapped_compact)

    return None, None


def is_invoice_detail_row(ws, row_idx: int, article_col: int, size_col: Optional[int]) -> bool:
    size_value = ws.cell(row_idx, size_col).value if size_col else None
    color_col = size_col - 1 if size_col and size_col > article_col + 1 else None
    color_value = ws.cell(row_idx, color_col).value if color_col else None
    has_variant = bool(norm_text(size_value) or norm_text(color_value))
    has_number = any(
        isinstance(ws.cell(row_idx, col).value, (int, float))
        for col in range(article_col + 1, ws.max_column + 1)
    )
    return has_variant and has_number


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format


def find_or_create_hs_col(ws, header_row: int, article_col: int, preferred_col: Optional[int]) -> int:
    if preferred_col:
        return preferred_col

    col = ws.max_column + 1
    ws.cell(header_row, col).value = "HS code (artikellijst)"

    for row_idx in range(1, ws.max_row + 1):
        copy_cell_style(ws.cell(row_idx, col - 1), ws.cell(row_idx, col))

    article_letter = ws.cell(header_row, article_col).column_letter
    hs_letter = ws.cell(header_row, col).column_letter
    if article_letter in ws.column_dimensions:
        ws.column_dimensions[hs_letter].width = ws.column_dimensions[article_letter].width
    return col


def fill_invoice(input_xlsx: Path, output_xlsx: Path, mapping_csv: Path) -> dict:
    mapping = load_mapping(mapping_csv)
    wb = load_workbook(input_xlsx)
    total_filled = 0
    unmatched = []

    invoice_sheets = [
        ws
        for ws in wb.worksheets
        if (
            ws.title.strip().lower() in {"invoice", "ci", "commercial invoice", "commerical invoice"}
            or "invoice" in ws.title.strip().lower()
        )
        and not re.search(r"\b(?:pl|packing)\b|packing\s*list", ws.title.strip(), flags=re.I)
    ]
    worksheets = invoice_sheets or wb.worksheets

    for ws in worksheets:
        header_row, article_col, hs_col, size_col = find_headers(ws)
        if not header_row or not article_col:
            continue
        created_hs_col = hs_col is None
        hs_col = find_or_create_hs_col(ws, header_row, article_col, hs_col)
        if created_hs_col and size_col and size_col >= hs_col:
            size_col += 1

        current_article_value = None

        for r in range(header_row + 1, ws.max_row + 1):
            row_text = norm_key(" ".join(norm_text(ws.cell(r, col).value) for col in range(1, ws.max_column + 1)))
            if "HS INFORMATION" in row_text or "H S INFORMATION" in row_text or "HSINFORMATION" in row_text:
                break

            article_value = ws.cell(r, article_col).value
            article = norm_key(article_value)
            if norm_header(article_value) in ARTICLE_HEADERS:
                current_article_value = None
                continue

            if article and article.startswith(("SUB TOTAL", "TOTAL", "EURO ", "N W", "NW", "G W", "GW", "NET WEIGHT", "GROSS WEIGHT")):
                current_article_value = None
                continue

            if article and not article.startswith("="):
                current_article_value = article_value
            elif not is_invoice_detail_row(ws, r, article_col, size_col):
                continue

            lookup_article = current_article_value
            if not norm_key(lookup_article):
                continue

            size_value = ws.cell(r, size_col).value if size_col else None
            matched_article, hs = lookup_article_from_excel_cell(mapping, lookup_article, size_value)
            if hs:
                ws.cell(r, hs_col).value = hs
                total_filled += 1
            elif is_invoice_detail_row(ws, r, article_col, size_col):
                unmatched.append({"sheet": ws.title, "row": r, "article": matched_article or lookup_article})

    wb.save(output_xlsx)
    return {"filled": total_filled, "unmatched": unmatched[:100], "unmatched_count": len(unmatched)}


def extract_pdf_words(page) -> List[dict]:
    words = []

    def visitor_text(text, _cm, tm, _font_dict, font_size):
        if not norm_text(text):
            return
        x = float(tm[4])
        y = float(tm[5])
        char_width = max(float(font_size or 10) * 0.5, 3)
        for match in re.finditer(r"\S+", text):
            words.append({"text": match.group(0), "x": x + (match.start() * char_width), "source_x": x, "y": y})

    page.extract_text(visitor_text=visitor_text)
    return words


def group_pdf_lines(words: List[dict], y_tolerance: float = 4.0) -> List[dict]:
    lines = []
    for word in sorted(words, key=lambda item: (-item["y"], item["x"])):
        for line in lines:
            if abs(line["y"] - word["y"]) <= y_tolerance:
                line["words"].append(word)
                line["y"] = (line["y"] + word["y"]) / 2
                break
        else:
            lines.append({"y": word["y"], "words": [word]})

    for line in lines:
        line["words"].sort(key=lambda item: item["x"])
        line["text"] = " ".join(word["text"] for word in line["words"])
    return lines


def find_pdf_article_column(lines: List[dict]) -> Optional[dict]:
    for line in lines:
        words = line["words"]
        tokens = [norm_pdf_token(word["text"]) for word in words]
        for idx, token in enumerate(tokens[:-1]):
            if token != "article" or tokens[idx + 1] not in {"no", "number"}:
                continue

            article_x = words[idx]["x"]
            header_end_x = words[idx + 1]["x"] + max(len(words[idx + 1]["text"]) * 6, 15)
            header_starts = [word["x"] for word in words]
            previous_starts = [x for x in header_starts if x < article_x - 10]
            next_starts = [x for x in header_starts if x > header_end_x + 10]
            left = ((max(previous_starts) + article_x) / 2) if previous_starts else article_x - 20
            right = ((header_end_x + min(next_starts)) / 2) if next_starts else header_end_x + 120
            return {"header_y": line["y"], "left": left, "right": right}
    return None


def find_article_in_text(value: str, mapping: Dict[str, str], patterns: List[Tuple[str, re.Pattern]]) -> Optional[str]:
    normalized = norm_key(value)
    if not normalized:
        return None
    if normalized in mapping:
        return normalized
    compact = compact_key(normalized)
    if compact in mapping:
        return compact

    for article, pattern in patterns:
        if pattern.search(normalized):
            return article
    return None


def find_article_prefix(value: str, mapping: Dict[str, str]) -> Optional[str]:
    compact = compact_key(value)
    if not compact:
        return None

    for article in sorted(mapping, key=len, reverse=True):
        if len(article) < 4 or not compact.startswith(article):
            continue
        if article.isdigit() and len(compact) > len(article) and compact[len(article)].isdigit():
            continue
        if re.fullmatch(r"\d{6,12}", compact):
            continue
        if re.fullmatch(r"\d+(\.\d+)?", norm_text(value)):
            continue
        return article
    return None


def display_article_code(value: str, lookup_article: str) -> str:
    lookup = norm_key(lookup_article)
    if not lookup:
        return lookup_article

    candidates = re.findall(
        r"\b\d{3,5}(?:\s+[A-Z]{1,6})?(?:\s+\d{1,3}(?:-\d{1,3})?)?\b",
        norm_key(value),
        flags=re.I,
    )
    for candidate in candidates:
        candidate = norm_key(candidate)
        if len(candidate) > len(lookup) and compact_key(candidate).startswith(compact_key(lookup)):
            return candidate
    return lookup_article


def extract_article_before_po(text: str, mapping: Dict[str, str]) -> Optional[str]:
    before_po = re.split(r"\bPO\s*#", text, flags=re.I)[0]
    candidates = re.findall(r"\b[A-Z]?\d{3,5}(?:\s*[A-Z]{2,6})?(?:\s*\d{1,3}(?:-\d{1,3})?)?\b", before_po, flags=re.I)
    for candidate in reversed(candidates):
        article = find_article_in_text(candidate, mapping, [])
        if article:
            return article
    return None


def remove_po_number(text: str) -> str:
    return norm_text(re.sub(r"\bPO\s*#\s*[A-Z0-9-]+", "", text, flags=re.I))


def extract_pdf_articles_from_po_lines(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    seen = set()

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        previous_line = ""
        for line_number, line in enumerate(text.splitlines(), start=1):
            clean_line = norm_text(line)
            if not clean_line:
                continue

            if re.search(r"\bPO\s*#", clean_line, flags=re.I):
                combined = norm_text(f"{previous_line} {clean_line}")
                article = extract_article_before_po(combined, mapping)
                if article:
                    key = (page_number, article)
                    if key not in seen:
                        seen.add(key)
                        rows.append(
                            {
                                "page": page_number,
                                "line": line_number,
                                "article": display_article_code(combined, article),
                                "lookup_article": article,
                                "hs_code": mapping[article],
                                "text": remove_po_number(combined),
                            }
                        )

            previous_line = clean_line

    return rows


def extract_pdf_articles_from_equi_style(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    line_pattern = re.compile(
        r"^\s*(?P<article>\d{3,5}\s+[A-Z]{1,6}\s+\d{1,3}[A-Z]?)\s+"
        r"(?P<quantity>\d+)\s+.+?\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+"
        r"(?P<amount>\d+(?:,\d{3})*(?:\.\d+)?)\s*$"
    )

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            match = line_pattern.match(line)
            if not match:
                continue

            article = norm_key(match.group("article"))
            hs = mapping.get(article) or mapping.get(compact_key(article))
            if not hs:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": match.group("quantity"),
                    "unit_price": match.group("unit_price"),
                    "amount": match.group("amount"),
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_shipment_columns(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []
    seen = set()

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"QUANTITY\s*\(PCS\).*UNIT\s+PRICE.*AMOUNT", full_text, flags=re.I | re.S):
            return []

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text) and not re.search(r"Our\s+Product\s+Code\s+Party'?s\s+Code\s+Order\s+No", text, flags=re.I):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            order_words = [
                word
                for word in words
                if re.fullmatch(r"\d{6}", word["text"])
                and float(word["x0"]) < 70
            ]

            for order_word in order_words:
                top = float(order_word["top"])
                row_words = [
                    word
                    for word in words
                    if top - 8 <= float(word["top"]) <= top + 12
                ]

                def column_text(left: float, right: float) -> str:
                    selected = [
                        word
                        for word in row_words
                        if left <= float(word["x0"]) <= right
                    ]
                    selected.sort(key=lambda word: float(word["x0"]))
                    return norm_text(" ".join(word["text"] for word in selected))

                article = norm_key(column_text(90, 150))
                quantity = norm_text(column_text(300, 360))
                unit_price = norm_text(column_text(370, 430))
                amount = norm_text(column_text(435, 540))

                if not re.fullmatch(r"[A-Z]?\d{3,5}\s+[A-Z]{1,6}\s+[A-Z0-9.]+", article):
                    continue

                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                key = (page_number, round(top), article)
                if key in seen:
                    continue
                seen.add(key)

                full_row_text = " ".join(
                    word["text"]
                    for word in sorted(row_words, key=lambda word: (float(word["x0"]), float(word["top"])))
                )
                rows.append(
                    {
                        "page": page_number,
                        "line": int(top),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": re.sub(r"\D", "", quantity),
                        "unit_price": unit_price,
                        "amount": amount,
                        "raw_text": full_row_text,
                        "text": norm_text(full_row_text),
                    }
                )

    return rows


def extract_pdf_articles_from_tarun_thermoware(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"Tarun\s+Thermoware", full_text, flags=re.I):
            return []
        if not re.search(r"Our\s+Product\s+Code\s+Party'?s\s+Code", full_text, flags=re.I):
            return []

        total_re = re.compile(
            r"(?P<total_quantity>\d+)\s*(?:Pair|Pcs\.?)\s+"
            r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
            r"(?P<amount>\d[\d,]*(?:\.\d{2})?)",
            flags=re.I,
        )

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text) and not re.search(r"Our\s+Product\s+Code\s+Party'?s\s+Code\s+Order\s+No", text, flags=re.I):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            page_rows = []
            for index, line in enumerate(lines):
                line_words = sorted(line["words"], key=lambda word: word["x"])
                text_line = norm_text(" ".join(word["text"] for word in line_words))
                article_part = norm_text(" ".join(word["text"] for word in line_words if 120 <= word["x"] <= 180))
                if not article_part:
                    continue

                order_part = norm_text(" ".join(word["text"] for word in line_words if 180 <= word["x"] <= 230))
                if not re.fullmatch(r"\d{6}", order_part):
                    continue

                article = find_article_in_text(article_part, mapping, [])
                if not article and index + 1 < len(lines):
                    next_words = sorted(lines[index + 1]["words"], key=lambda word: word["x"])
                    next_left = norm_text(" ".join(word["text"] for word in next_words if 120 <= word["x"] <= 180))
                    next_order = norm_text(" ".join(word["text"] for word in next_words if 180 <= word["x"] <= 230))
                    if next_left and not next_order:
                        article = find_article_in_text(f"{article_part} {next_left}", mapping, [])
                        if article:
                            text_line = norm_text(f"{text_line} {next_left}")
                            article_part = norm_text(f"{article_part} {next_left}")
                if not article:
                    continue

                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                qty_part = norm_text(" ".join(word["text"] for word in line_words if 400 <= word["x"] <= 430))
                qty_match = re.search(r"\d+", qty_part)
                if not qty_match:
                    continue

                total_text = norm_text(" ".join(word["text"] for word in line_words if word["x"] >= 430))
                total_match = total_re.search(total_text)
                row = {
                    "page": page_number,
                    "line": int(-line["y"]),
                    "article": canonical_article_key(article, mapping),
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": qty_match.group(0),
                    "unit_price": "",
                    "amount": "",
                    "skip_value_fallback": True,
                    "raw_text": text_line,
                    "text": text_line,
                }
                page_rows.append(row)

                if total_match:
                    rate = parse_pdf_number(total_match.group("unit_price"))
                    if isinstance(rate, (int, float)):
                        for pending in reversed(page_rows):
                            if pending.get("unit_price"):
                                break
                            quantity = parse_pdf_number(pending.get("quantity", ""))
                            if isinstance(quantity, (int, float)):
                                pending["unit_price"] = total_match.group("unit_price")
                                pending["amount"] = round(float(quantity) * float(rate), 2)

            rows.extend(page_rows)

    return rows


def extract_pdf_articles_from_ibrahim_buyer_code(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"IBRAHIM\s+INTERNATIONAL", full_text, flags=re.I):
            return []
        if not re.search(r"Our\s+Code\s+&\s+Name\s+Buyer\s+Code", full_text, flags=re.I):
            return []

        buyer_code_re = re.compile(
            r"(?P<article>\d{3,5}\s+[A-Z]{1,6}(?:\s+\d+)?)\s+\d{6}\b",
            flags=re.I,
        )
        total_re = re.compile(
            r"(?P<total_quantity>\d+)\s*(?:Pcs|Pair|Prs|Set)\s+"
            r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
            r"(?P<amount>\d[\d,]*(?:\.\d{2})?)",
            flags=re.I,
        )

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text) and not re.search(r"Our\s+Code\s+&\s+Name\s+Buyer\s+Code", text, flags=re.I):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            page_rows = []
            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                article_match = buyer_code_re.search(text_line)
                if not article_match:
                    continue

                article = norm_key(article_match.group("article"))
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                after_article = text_line[article_match.end() :]
                total_match = total_re.search(after_article)
                quantity_text = after_article[: total_match.start()] if total_match else after_article
                numbers_before_total = re.findall(r"\d+(?:[,.]\d+)?", quantity_text)
                if not numbers_before_total:
                    continue

                row = {
                    "page": page_number,
                    "line": int(-line["y"]),
                    "article": canonical_article_key(article, mapping),
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": numbers_before_total[-1],
                    "unit_price": "",
                    "amount": "",
                    "skip_value_fallback": True,
                    "raw_text": text_line,
                    "text": text_line,
                }
                page_rows.append(row)

                if total_match:
                    rate = parse_pdf_number(total_match.group("unit_price"))
                    if isinstance(rate, (int, float)):
                        for pending in reversed(page_rows):
                            if pending.get("unit_price"):
                                break
                            quantity = parse_pdf_number(pending.get("quantity", ""))
                            if isinstance(quantity, (int, float)):
                                pending["unit_price"] = total_match.group("unit_price")
                                pending["amount"] = round(float(quantity) * float(rate), 2)

            rows.extend(page_rows)

    return rows


def extract_pdf_articles_from_ruksh_rows(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"RUKSH\s+ENTERPRISES", full_text, flags=re.I):
            return []
        if not re.search(r"BUYER\s+ARTICLE.*STYLE\s+DESCRIPTION.*Rate.*Amount", full_text, flags=re.I | re.S):
            return []

        row_re = re.compile(
            r"^(?P<order>\d{6})\s+"
            r"(?P<article>\d{3,5}\s+[A-Z]{1,6})\s+"
            r"(?P<invoice_hs>\d{6,10})\s+"
            r".+?\s+"
            r"(?P<color>[A-Z]+)\s+"
            r"(?P<size>[A-Z0-9./-]+)\s+"
            r"(?P<quantity>\d+(?:[,.]\d+)?)\s+"
            r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
            r"(?P<amount>\d[\d,]*(?:\.\d{2})?)$",
            flags=re.I,
        )

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                match = row_re.match(text_line)
                if not match:
                    continue

                article = canonical_article_key(match.group("article"), mapping)
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                rows.append(
                    {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": match.group("quantity"),
                        "unit_price": match.group("unit_price"),
                        "amount": match.group("amount"),
                        "skip_value_fallback": True,
                        "raw_text": text_line,
                        "text": text_line,
                    }
                )

    return rows


def extract_pdf_articles_from_mark_equestrian_party_code(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"MARK\s+EQUESTRIAN", full_text, flags=re.I):
            return []
        if not re.search(r"Our\s+Product\s+Code\s+Party'?s\s+Code\s+Order\s+No", full_text, flags=re.I):
            return []

        party_code_re = re.compile(
            r"(?P<article>\d{3,5}\s+[A-Z]{2,6}\s+\d{2,3})\s+502516\b",
            flags=re.I,
        )
        quantity_re = re.compile(r"\b(?:Child|child)\s+(?P<quantity>\d+)(?:\s+\d+\s*Pcs|\s*\d*Pcs)?\b")

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text) and not re.search(r"Our\s+Product\s+Code\s+Party'?s\s+Code\s+Order\s+No", text, flags=re.I):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                article_match = party_code_re.search(text_line)
                if not article_match:
                    continue

                article = norm_key(article_match.group("article"))
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                quantity = ""
                quantity_match = quantity_re.search(text_line[article_match.end() :])
                if quantity_match:
                    quantity = quantity_match.group("quantity")

                rows.append(
                    {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": quantity,
                        "skip_value_fallback": True,
                        "raw_text": text_line,
                        "text": text_line,
                    }
                )

    return rows


def extract_pdf_articles_from_leather_art_variants(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    def variant_size_code(value: str) -> str:
        text = norm_key(value)
        text = re.sub(r"\b(CM|CMS|PONY|COB|FULL|XFULL|X-FULL|EX-FULL|EXFULL|SHET|SHETLAND)\b", " ", text)
        text = re.sub(r"[^A-Z0-9-]+", " ", text).strip()
        parts = text.split()
        return parts[0] if parts else ""

    def lookup_variant(base_article: str, color: str, size: str) -> Tuple[str, Optional[str]]:
        base_article = norm_key(base_article)
        color = norm_key(color)
        size = norm_key(size)
        candidates = []
        if base_article in mapping or compact_key(base_article) in mapping:
            candidates.append(base_article)
        if color:
            if size:
                candidates.insert(0, norm_key(f"{base_article} {color} {size}"))
            candidates.append(norm_key(f"{base_article} {color}"))
        candidates.append(base_article)

        for candidate in candidates:
            hs = mapping.get(candidate) or mapping.get(compact_key(candidate))
            if hs:
                return candidate, hs
        return base_article, None

    def flush_current(current: Optional[dict], variants: List[str]) -> None:
        if not current:
            return

        rate = float(parse_pdf_number(current["unit_price"]) or 0)
        emitted = False
        for variant in variants:
            variant_match = re.match(r"^\((?P<color>[A-Z]+)\s+(?P<body>.+)\)$", norm_text(variant), flags=re.I)
            if not variant_match:
                continue

            color = variant_match.group("color")
            for part in variant_match.group("body").split(","):
                part = norm_text(part)
                qty_match = re.search(r"/\s*(?P<quantity>\d+)\b", part)
                if not qty_match:
                    continue

                quantity = qty_match.group("quantity")
                size = variant_size_code(part[: qty_match.start()])
                article, hs = lookup_variant(current["base_article"], color, size)
                if not hs:
                    continue

                amount = float(quantity) * rate if rate else ""
                rows.append(
                    {
                        "page": current["page"],
                        "line": current["line"],
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": quantity,
                        "unit_price": current["unit_price"],
                        "amount": f"{amount:.2f}" if amount != "" else "",
                        "skip_value_fallback": True,
                        "raw_text": norm_text(f"{current['raw_text']} {variant}"),
                        "text": norm_text(f"{current['raw_text']} {variant}"),
                    }
                )
                emitted = True

        if emitted:
            return

        article, hs = lookup_variant(current["base_article"], "", "")
        if not hs:
            return
        rows.append(
            {
                "page": current["page"],
                "line": current["line"],
                "article": article,
                "lookup_article": article,
                "hs_code": hs,
                "quantity": current["quantity"],
                "unit_price": current["unit_price"],
                "amount": current["amount"],
                "skip_value_fallback": True,
                "raw_text": current["raw_text"],
                "text": current["raw_text"],
            }
        )

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"\bLEATHER\s+ART\b", full_text, flags=re.I):
            return []
        if not re.search(r"PARTS\s+OF\s+(?:NON-)?LEATHER\s+HARNESS\s+GOODS", full_text, flags=re.I):
            return []

        product_re = re.compile(
            r"^\s*(?P<serial>\d+)\.\s+"
            r"(?P<article>[A-Z]?\d{3,5}(?:\s+[A-Z]{1,6}(?:\s+\d+(?:-\d+)?)?)?)\s+"
            r"(?P<quantity>\d+)\s+(?:PCS|PRS)\.?\s+"
            r"(?P<unit_price>\d+(?:\.\d+)?)\s+"
            r"(?P<amount>\d[\d,]*(?:\.\d{2})?)",
            flags=re.I,
        )
        variant_re = re.compile(r"^\([A-Z]+\s+.+/\d+.*\)$", flags=re.I)
        stop_re = re.compile(r"^(Group Total|Invoice Total|Grand Total|TOTAL CARTON)", flags=re.I)

        current = None
        variants: List[str] = []

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                if not text_line:
                    continue

                product_match = product_re.match(text_line)
                if product_match:
                    flush_current(current, variants)
                    current = {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "base_article": norm_key(product_match.group("article")),
                        "quantity": product_match.group("quantity"),
                        "unit_price": product_match.group("unit_price"),
                        "amount": product_match.group("amount"),
                        "raw_text": text_line,
                    }
                    variants = []
                    continue

                if current and variant_re.match(text_line):
                    variants.append(text_line)
                    continue

                if current and stop_re.match(text_line):
                    flush_current(current, variants)
                    current = None
                    variants = []

        flush_current(current, variants)

    return rows


def extract_pdf_articles_from_gng_pet_rows(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"\bGNG\s+PET\b", full_text, flags=re.I):
        return []
    if not re.search(r"Art\.?\s*No\.?\s+Item\s+Order\s+No\.?\s+Size\s+Colour\s+Pcs\s+Rate\s+Amount", full_text, flags=re.I):
        return []

    rows = []
    money = r"\d[\d,]*(?:\.\d{2})|\d[\d.]*(?:,\d{2})"
    line_re = re.compile(
        r"^(?P<article>[A-Z]?\d{3,5}\s+[A-Z]{2,6}\s+[A-Z0-9]+)\s+.+?\s+"
        r"(?P<order>\d{6})\s+.+?\s+"
        r"(?P<quantity>\d+)\s+"
        rf"(?P<unit_price>{money})\s+"
        rf"(?P<amount>{money})\s*€?\s*$",
        flags=re.I,
    )

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line).replace("\ufffd", "").strip()
            match = line_re.match(text_line)
            if not match:
                continue

            article = norm_key(match.group("article"))
            hs = mapping.get(article) or mapping.get(compact_key(article))
            if not hs:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": match.group("quantity"),
                    "unit_price": match.group("unit_price"),
                    "amount": match.group("amount"),
                    "skip_value_fallback": True,
                    "raw_text": text_line,
                    "text": text_line,
                }
            )

    return rows


def extract_pdf_articles_from_zhongshan_biaoqi(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"ZHONGSHAN\s+BIAOQI\s+HOUSEWARE", full_text, flags=re.I):
        return []
    if not re.search(r"Item\s+No\.?\s+Production\s+Description", full_text, flags=re.I):
        return []

    value_re = re.compile(
        r"^(?P<unit_price>\d[\d,]*(?:\.\d{2})?)\s+"
        r"(?P<quantity>\d[\d,]*)\s+"
        r"(?P<amount>\d[\d,]*(?:\.\d{2})?)\s+"
        r"(?P<cartons>\d+)\s*$"
    )
    article_re = re.compile(r"^(?P<article>[A-Z]?\d{3,5}\s+[A-Z]{2,6}\s+[A-Z0-9]+)\b(?P<desc>.*)$", flags=re.I)

    rows = []
    values: List[dict] = []
    articles: List[dict] = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line).strip()
            value_match = value_re.match(text_line)
            if value_match:
                values.append(value_match.groupdict())
                continue

            article_match = article_re.match(text_line)
            if not article_match:
                continue

            article = norm_key(article_match.group("article"))
            hs = mapping.get(article) or mapping.get(compact_key(article))
            if not hs:
                continue

            articles.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "raw_text": text_line,
                    "text": text_line,
                }
            )

    for article_row, value_row in zip(articles, values):
        row = dict(article_row)
        row["quantity"] = value_row["quantity"].replace(",", "")
        row["unit_price"] = value_row["unit_price"]
        row["amount"] = value_row["amount"]
        row["skip_value_fallback"] = True
        rows.append(row)

    return rows


def extract_pdf_articles_from_changzhou_biseen(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"CHANGZHOU\s+BISEEN\s+HARNESS\s+EQUIPMENT", full_text, flags=re.I):
        return []
    if not re.search(r"style\s+no", full_text, flags=re.I) or not re.search(r"qty", full_text, flags=re.I):
        return []

    money = r"\$?\d[\d.,]*"
    base_re = re.compile(r"^(?P<base>\d{3,5}\s+[A-Z]{2,6})\s*$", flags=re.I)
    variant_re = re.compile(r"^(?P<variant>[A-Z0-9]{1,5})\s*$", flags=re.I)
    inline_article_re = re.compile(
        r"^(?P<article>\d{3,5}\s+[A-Z]{2,6}\s+[A-Z0-9]+)\b\s*(?P<rest>.*)$",
        flags=re.I,
    )
    value_re = re.compile(
        rf"^(?P<desc>.+?)\s+(?P<quantity>\d+)\s+"
        rf"(?P<unit_price>{money})\s+(?P<amount>{money})\s*$",
        flags=re.I,
    )

    rows = []
    pending_base = ""
    current_article = ""
    current_start = {"page": 0, "line": 0}
    desc_lines: List[str] = []

    def add_row(page_number: int, line_number: int, article_text: str, value_match, raw_text: str) -> None:
        article = norm_key(article_text)
        hs = mapping.get(article) or mapping.get(compact_key(article))
        if not hs:
            return
        rows.append(
            {
                "page": current_start.get("page") or page_number,
                "line": current_start.get("line") or line_number,
                "article": article,
                "lookup_article": article,
                "hs_code": hs,
                "quantity": value_match.group("quantity"),
                "unit_price": value_match.group("unit_price").replace("$", ""),
                "amount": value_match.group("amount").replace("$", ""),
                "skip_value_fallback": True,
                "raw_text": raw_text,
                "text": raw_text,
            }
        )

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line).strip()
            if not text_line:
                continue
            if re.match(r"^total\b", text_line, flags=re.I):
                current_article = ""
                pending_base = ""
                desc_lines = []
                continue

            base_match = base_re.match(text_line)
            if base_match:
                pending_base = base_match.group("base")
                current_article = ""
                desc_lines = []
                current_start = {"page": page_number, "line": line_number}
                continue

            if pending_base:
                variant_match = variant_re.match(text_line)
                if variant_match:
                    current_article = f"{pending_base} {variant_match.group('variant')}"
                    pending_base = ""
                    desc_lines = []
                    continue

            inline_match = inline_article_re.match(text_line)
            if inline_match:
                current_article = inline_match.group("article")
                current_start = {"page": page_number, "line": line_number}
                rest = inline_match.group("rest").strip()
                desc_lines = []
                if rest:
                    value_match = value_re.match(rest)
                    if value_match:
                        add_row(page_number, line_number, current_article, value_match, text_line)
                        current_article = ""
                    else:
                        desc_lines.append(rest)
                continue

            if not current_article:
                continue

            combined_text = " ".join(desc_lines + [text_line]).strip()
            value_match = value_re.match(combined_text)
            if value_match:
                raw_text = f"{current_article} {combined_text}"
                add_row(page_number, line_number, current_article, value_match, raw_text)
                current_article = ""
                desc_lines = []
            else:
                desc_lines.append(text_line)

    return rows


def extract_pdf_articles_from_quanzhou_xingye(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"QUANZHOU\s+XINGYE\s+TRAVELLING\s+PRODUCTS", full_text, flags=re.I):
        return []
    if not re.search(r"ORDER\s+NUMBER\s+ART\.?NAME.*QUANTITY.*UNIT\s+PRICE.*AMOUNT", full_text, flags=re.I | re.S):
        return []

    start_re = re.compile(
        r"^(?P<order>\d{6})\s+(?P<article>[A-Z]?\d{3,5}\s+[A-Z]{1,6}\s+[A-Z0-9]+)\b(?P<rest>.*)$",
        flags=re.I,
    )
    value_re = re.compile(r"\b(?P<cartons>\d+)\s+(?P<quantity>\d+)\s+USD\s*(?P<unit_price>\d[\d.,]*)\s*$", flags=re.I)
    amount_re = re.compile(r"^US\$\s*(?P<amount>\d[\d,]*(?:\.\d{2})?)\s*$", flags=re.I)

    rows = []
    pending: Optional[dict] = None

    def flush_with_amount(amount: str) -> None:
        nonlocal pending
        if not pending:
            return
        article = norm_key(pending["article"])
        hs = mapping.get(article) or mapping.get(compact_key(article))
        if hs:
            raw_text = norm_text(" ".join(pending.get("parts", [])))
            rows.append(
                {
                    "page": pending["page"],
                    "line": pending["line"],
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": pending.get("quantity", ""),
                    "unit_price": pending.get("unit_price", ""),
                    "amount": amount,
                    "skip_value_fallback": True,
                    "raw_text": raw_text,
                    "text": raw_text,
                }
            )
        pending = None

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line).strip()
            if not text_line:
                continue
            if re.match(r"^(TOTAL|20%|80%)\b", text_line, flags=re.I):
                pending = None
                continue

            amount_match = amount_re.match(text_line)
            if amount_match:
                flush_with_amount(amount_match.group("amount"))
                continue

            start_match = start_re.match(text_line)
            if start_match:
                pending = {
                    "page": page_number,
                    "line": line_number,
                    "article": start_match.group("article"),
                    "parts": [text_line],
                    "quantity": "",
                    "unit_price": "",
                }
                rest = start_match.group("rest").strip()
                value_match = value_re.search(rest)
                if value_match:
                    pending["quantity"] = value_match.group("quantity")
                    pending["unit_price"] = value_match.group("unit_price")
                continue

            if not pending:
                continue

            pending["parts"].append(text_line)
            value_match = value_re.search(text_line)
            if value_match:
                pending["quantity"] = value_match.group("quantity")
                pending["unit_price"] = value_match.group("unit_price")

    return rows


def extract_pdf_articles_from_xiamen_grassland(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"XIAMEN\s+GRASSLAND\s+SADDLERY\s+SPORT", full_text, flags=re.I):
        return []
    if not re.search(r"Barcode\s+Item\s+No", full_text, flags=re.I):
        return []

    article_re = re.compile(
        r"^(?P<barcode>\d{10,14})\s+"
        r"(?P<article>\d{3,5}\s+[A-Z]{2,6}\s+[A-Z0-9]+)\b(?P<rest>.*)$",
        flags=re.I,
    )
    value_re = re.compile(
        r"\b(?P<quantity>\d+)\s+\$(?P<unit_price>\d[\d.,]*)\s+US\$(?P<amount>\d[\d,]*(?:\.\d{2})?)\s*$",
        flags=re.I,
    )

    rows = []
    pending: Optional[dict] = None

    def append_row(page_number: int, line_number: int, pending_row: dict, value_match) -> None:
        article = norm_key(pending_row["article"])
        hs = mapping.get(article) or mapping.get(compact_key(article))
        if not hs:
            return
        raw_text = norm_text(" ".join(pending_row["parts"]))
        rows.append(
            {
                "page": pending_row["page"],
                "line": pending_row["line"],
                "article": article,
                "lookup_article": article,
                "hs_code": hs,
                "quantity": value_match.group("quantity"),
                "unit_price": value_match.group("unit_price"),
                "amount": value_match.group("amount"),
                "skip_value_fallback": True,
                "raw_text": raw_text,
                "text": raw_text,
            }
        )

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line).strip()
            if not text_line:
                continue
            if re.match(r"^(TOTAL\s+AMOUNT|SAY\s+TOTAL|Balance\b)", text_line, flags=re.I):
                pending = None
                continue

            article_match = article_re.match(text_line)
            if article_match:
                pending = {
                    "page": page_number,
                    "line": line_number,
                    "article": article_match.group("article"),
                    "parts": [text_line],
                }
                value_match = value_re.search(article_match.group("rest"))
                if value_match:
                    append_row(page_number, line_number, pending, value_match)
                    pending = None
                continue

            if not pending:
                continue

            pending["parts"].append(text_line)
            combined = " ".join(pending["parts"])
            value_match = value_re.search(combined)
            if value_match:
                append_row(page_number, line_number, pending, value_match)
                pending = None

    return rows


def extract_pdf_articles_from_jone_shou_scan(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pypdfium2 as pdfium
        ocr = create_rapidocr_engine()
    except ImportError:
        return []

    compact_to_article: Dict[str, str] = {}
    for article in mapping:
        compact = compact_key(article)
        if len(compact) < 6 or compact.isdigit() or not re.search(r"[A-Z]", compact):
            continue
        current = compact_to_article.get(compact)
        if not current or (" " in article and " " not in current):
            compact_to_article[compact] = article

    def ocr_article(text: str) -> Optional[str]:
        compact = compact_key(text)
        if not compact:
            return None
        variants = [compact]
        if re.search(r"\d+C[NX]$", compact):
            variants.append(re.sub(r"C[NX]$", "CM", compact))
        if "ZV" in compact:
            variants.append(compact.replace("ZV", "ZW"))
        for variant in variants:
            article = compact_to_article.get(variant)
            if article:
                return article
        for mapped_compact in sorted(compact_to_article, key=len, reverse=True):
            if len(mapped_compact) >= 7 and mapped_compact in compact:
                return compact_to_article[mapped_compact]
        return None

    def nearby_value(tokens: List[dict], article_y: float, left: float, right: float, pattern: str) -> Optional[dict]:
        candidates = [
            token
            for token in tokens
            if left <= token["x"] <= right
            and abs(token["y"] - article_y) <= 24
            and re.search(pattern, token["text"], flags=re.I)
        ]
        if not candidates:
            return None
        return min(candidates, key=lambda token: abs(token["y"] - article_y))

    rows = []
    seen = set()
    ocr_text_parts = []
    document = pdfium.PdfDocument(str(input_pdf))

    try:
        for page_number in range(len(document)):
            page = document[page_number]
            try:
                image = page.render(scale=2).to_pil()
            finally:
                if hasattr(page, "close"):
                    page.close()

            width, height = image.size
            table_image = image.crop(
                (
                    max(0, int(width * 0.08)),
                    int(height * 0.16),
                    min(width, int(width * 0.96)),
                    min(height, int(height * 0.94)),
                )
            )

            tokens = []
            for box, text, confidence in iter_rapidocr_results(ocr, table_image):
                if confidence < 0.75:
                    continue
                xs = [point[0] for point in box]
                ys = [point[1] for point in box]
                token = {"x": min(xs), "y": min(ys), "text": norm_text(text)}
                tokens.append(token)
                ocr_text_parts.append(token["text"])

            for token in tokens:
                if not 130 <= token["x"] <= 260:
                    continue
                article = ocr_article(token["text"])
                if not article:
                    continue

                qty_token = nearby_value(tokens, token["y"], 540, 670, r"\d+\s*(?:PCS|PAIRS?)\b")
                unit_token = nearby_value(tokens, token["y"], 700, 840, r"(?:USD)?\d+[\d.,]*")
                amount_token = nearby_value(tokens, token["y"], 880, 1020, r"(?:USD)?\d+[\d,]*(?:\.\d{2})?")
                if not (qty_token and unit_token and amount_token):
                    continue

                quantity = re.sub(r"\D", "", qty_token["text"])
                unit_price = re.sub(r"(?i)^USD", "", unit_token["text"]).strip()
                amount = re.sub(r"(?i)^USD", "", amount_token["text"]).strip()
                key = (page_number + 1, article, quantity, amount)
                if key in seen:
                    continue
                seen.add(key)

                raw_text = norm_text(f"{article} {qty_token['text']} {unit_token['text']} {amount_token['text']}")
                rows.append(
                    {
                        "page": page_number + 1,
                        "line": int(token["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": mapping[article] or mapping.get(compact_key(article), ""),
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "amount": amount,
                        "skip_value_fallback": True,
                        "raw_text": raw_text,
                        "text": raw_text,
                    }
                )
    finally:
        document.close()

    ocr_text = " ".join(ocr_text_parts)
    if not re.search(r"(?:MADE\s*IN\s*TAIWAN|FOB\s*KAOHSIUNG|JS260616040)", ocr_text, flags=re.I):
        return []
    return rows if len(rows) >= 10 else []


def extract_pdf_articles_from_changzhou_ziyuan_rows(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not re.search(r"CHANGZHOU\s+ZIYUAN\s+SPORTS", full_text, flags=re.I):
        return []
    if not re.search(r"ITEM\s+NO\.?\s+DESCRIPTIONS\s+QUANTITIES\s+FOB\s+SHANGHAI\s+AMOUNT", full_text, flags=re.I):
        return []

    rows = []
    line_re = re.compile(
        r"^(?:(?P<carton>[A-Z]{0,4}\d[A-Z0-9-]*)\s+)?"
        r"(?P<order>\d{6})\s+"
        r"(?P<item_text>.+?)\s+"
        r"(?P<quantity>\d+)\s+"
        r"US\$?(?P<unit_price>\d[\d,.]*)\s+"
        r"US\$?(?P<amount>\d[\d,.]*)$",
        flags=re.I,
    )

    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            text_line = norm_text(line)
            match = line_re.match(text_line)
            if not match:
                continue

            lookup_article = find_article_prefix(match.group("item_text"), mapping)
            if not lookup_article:
                continue

            hs = mapping.get(lookup_article) or mapping.get(compact_key(lookup_article))
            if not hs:
                continue

            article = canonical_article_key(lookup_article, mapping)
            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": lookup_article,
                    "hs_code": hs,
                    "quantity": match.group("quantity"),
                    "unit_price": match.group("unit_price"),
                    "amount": match.group("amount"),
                    "skip_value_fallback": True,
                    "raw_text": text_line,
                    "text": text_line,
                }
            )

    return rows


def extract_pdf_articles_from_karan_letex_rows(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"KARAN\s+LETEX\s+LTD", full_text, flags=re.I):
            return []
        if not re.search(r"Knitted\s+Man\s+Made\s+Breeches", full_text, flags=re.I):
            return []

        line_re = re.compile(
            r"Knitted\s+Man\s+Made\s+Breeches\s+-\s+(?:Ladies|Girls)\s+"
            r"(?P<article>\d{3,5}\s+[A-Z]{2,6}\s+\d{2,3})\s+"
            r"(?P<quantity>\d+)\s*Pcs\.?\s+"
            r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
            r"(?P<amount>\d[\d,]*(?:\.\d{2})?)",
            flags=re.I,
        )

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                match = line_re.search(text_line)
                if not match:
                    continue

                article = norm_key(match.group("article"))
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                rows.append(
                    {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": match.group("quantity"),
                        "unit_price": match.group("unit_price"),
                        "amount": match.group("amount"),
                        "skip_value_fallback": True,
                        "raw_text": text_line,
                        "text": text_line,
                    }
                )

    return rows


def canonical_article_key(article: str, mapping: Dict[str, str]) -> str:
    compact = compact_key(article)
    matches = [key for key in mapping if compact_key(key) == compact]
    spaced = [key for key in matches if " " in key]
    if spaced:
        return min(spaced, key=len)
    return matches[0] if matches else article

def extract_pdf_articles_from_mehra_shoes(
    input_pdf: Path,
    mapping: Dict[str, str],
) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows: List[dict] = []
    pending_rows: List[dict] = []
    seen = set()

    # Voorbeelden:
    # 7237BR37 105367 Brown 37 10.0
    # 7142ZW37 105367 Black 37 E.W - 20.0
    # 7100ZW0 105367 Black 47 - 200.0 200.0 Pairs 0.15 30.00
    article_re = re.compile(
        r"\b(?P<article>\d{3,5}\s*[A-Z]{1,6}\s*\d{0,3})\b"
        r".*?"
        r"(?:\b(?P<order>\d{6})\b.*?)?"
        r"(?:-\s*)?"
        r"(?P<quantity>\d+(?:[,.]\d+)?)"
        r"(?:\s+"
        r"(?P<group_quantity>\d+(?:[,.]\d+)?)\s*"
        r"(?:Pairs?|Pair|Prs|Pcs)\.?\s+"
        r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
        r"(?P<amount>\d[\d,]*(?:\.\d{2})?)"
        r")?\s*$",
        flags=re.I,
    )

    # Soms staat het groepstotaal als los herkend gedeelte aan het einde.
    total_re = re.compile(
        r"(?P<group_quantity>\d+(?:[,.]\d+)?)\s*"
        r"(?:Pairs?|Prs|Pcs)\.?\s+"
        r"(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
        r"(?P<amount>\d[\d,]*(?:\.\d{2})?)\s*$",
        flags=re.I,
    )

    def lookup_mehra_article(raw_article: str) -> Tuple[str, Optional[str]]:
        normalized = norm_key(raw_article)
        compact = compact_key(normalized)

        candidates = [
            normalized,
            compact,
        ]

        # Mehra gebruikt bijvoorbeeld 7237BR37 terwijl de mapping eventueel
        # als "7237 BR 37" is opgeslagen.
        spaced_match = re.fullmatch(
            r"(?P<number>\d{3,5})(?P<letters>[A-Z]{1,6})(?P<size>\d{0,3})",
            compact,
        )
        if spaced_match:
            number = spaced_match.group("number")
            letters = spaced_match.group("letters")
            size = spaced_match.group("size")

            candidates.extend(
                [
                    norm_key(f"{number} {letters} {size}") if size else norm_key(f"{number} {letters}"),
                    norm_key(f"{number}{letters}{size}"),
                    norm_key(f"{number} {letters}"),
                ]
            )

        for candidate in candidates:
            hs = mapping.get(candidate) or mapping.get(compact_key(candidate))
            if hs:
                article = canonical_article_key(candidate, mapping)
                return article, hs

        return normalized, None

    def flush_pending(unit_price_text: str) -> None:
        nonlocal pending_rows

        unit_price_number = parse_pdf_number(unit_price_text)
        if not isinstance(unit_price_number, (int, float)):
            pending_rows = []
            return

        for pending in pending_rows:
            quantity_number = parse_pdf_number(pending["quantity"])
            if not isinstance(quantity_number, (int, float)):
                continue

            amount = round(float(quantity_number) * float(unit_price_number), 2)

            pending["unit_price"] = f"{float(unit_price_number):.2f}"
            pending["amount"] = f"{amount:.2f}"
            rows.append(pending)

        pending_rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(
            page.extract_text(layout=True) or page.extract_text() or ""
            for page in pdf.pages
        )

        if not re.search(r"\bMEHRA\s+SHOES\b", full_text, flags=re.I):
            return []

        if not re.search(
            r"Our\s+Product\s+Code\s+Party'?s\s+Code\s+Order\s+No",
            full_text,
            flags=re.I,
        ):
            return []

        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(
                x_tolerance=2,
                y_tolerance=3,
                keep_blank_chars=False,
            )

            lines = group_pdf_lines(
                [
                    {
                        "text": word["text"],
                        "x": float(word["x0"]),
                        "source_x": float(word["x0"]),
                        "y": -float(word["top"]),
                    }
                    for word in words
                ],
                y_tolerance=3,
            )

            for line in lines:
                line_words = sorted(line["words"], key=lambda item: item["x"])
                text_line = norm_text(
                    " ".join(word["text"] for word in line_words)
                )

                if not text_line:
                    continue

                upper_line = text_line.upper()
                if upper_line.startswith(
                    (
                        "EXPORT INVOICE",
                        "OUR PRODUCT CODE",
                        "SUB TOTAL",
                        "GRAND TOT",
                        "TOTAL FOB",
                        "DECLARATION",
                        "SIGNATURE",
                    )
                ):
                    continue

                article_match = article_re.search(text_line)
                if not article_match:
                    continue

                raw_article = article_match.group("article")
                article, hs = lookup_mehra_article(raw_article)
                if not hs:
                    continue

                quantity_number = parse_pdf_number(article_match.group("quantity"))
                if not isinstance(quantity_number, (int, float)) or quantity_number <= 0:
                    continue

                # use the parsed numeric quantity (quantity_number) instead of the
                # raw regex group so downstream logic receives a numeric value
                # (int/float) rather than a string.
                quantity = quantity_number
                key = (
                    page_number,
                    int(-line["y"]),
                    compact_key(article),
                    quantity,
                )
                if key in seen:
                    continue
                seen.add(key)

                pending_rows.append(
                    {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": quantity,
                        "unit_price": "",
                        "amount": "",
                        "skip_value_fallback": True,
                        "raw_text": text_line,
                        "text": text_line,
                    }
                )

                unit_price = article_match.group("unit_price")

                if not unit_price:
                    total_match = total_re.search(text_line)
                    if total_match:
                        unit_price = total_match.group("unit_price")

                # De prijs aan het einde van de regel geldt voor alle
                # voorgaande maten binnen dezelfde productgroep.
                if unit_price:
                    flush_pending(unit_price)

    # Onvolledige regels zonder groepsprijs niet aan het resultaat toevoegen.
    return rows

def extract_pdf_articles_from_maharaja_scan(
    input_pdf: Path,
    reader: PdfReader,
    mapping: Dict[str, str],
) -> List[dict]:
    if any(norm_text(get_pdf_page_text(page)) for page in reader.pages):
        return []

    try:
        import pypdfium2 as pdfium
        ocr = create_rapidocr_engine()
    except ImportError as exc:
        raise RuntimeError(
            f"OCR kon niet worden gestart. Werkelijke fout: {exc}"
        ) from exc

    rows = []
    seen = set()
    document = pdfium.PdfDocument(str(input_pdf))
    amount_re = re.compile(r"\d[\d,]*(?:\.\d{2})")
    qty_re = re.compile(r"(?P<quantity>\d+)\s*(?:PCS|PRS|SET)\b", flags=re.I)
    compact_articles = sorted(
        [(article, compact_key(article)) for article in mapping],
        key=lambda item: len(item[1]),
        reverse=True,
    )

    def normalize_maharaja_ocr_text(value: str) -> str:
        text = norm_text(value).replace("'", "")
        text = re.sub(r"\b[Ss](\d{3})\b", r"5\1", text)
        text = re.sub(r"\b31[Oo0][Oo0]\b", "3100", text)
        text = re.sub(r"\b(\d{3,5})\s+B0\b", r"\1 BO", text, flags=re.I)
        text = re.sub(r"\b(\d{3,5})\s+[2Zz][Ww]\b", r"\1 ZW", text)
        text = re.sub(r"\b(\d{3,5})\s+ZW\s+ZW\b", r"\1 ZWZW", text, flags=re.I)
        text = re.sub(r"\b(\d{3,5})\s+2W\s+ZW\b", r"\1 ZWZW", text, flags=re.I)
        text = re.sub(r"\b(\d{3,5})\s+ZW\s+2W\b", r"\1 ZWZW", text, flags=re.I)
        text = re.sub(r"\b([2Zz][Ww])\s+([A-Z0-9])", r"ZW \2", text)
        text = re.sub(r"\b(\d{3,5}\s+[A-Z]{1,6})\s+O\b", r"\1 0", text, flags=re.I)
        return norm_text(text)

    def find_compact_article(value: str) -> Optional[str]:
        compact = compact_key(normalize_maharaja_ocr_text(value))
        if not compact:
            return None
        for article, article_compact in compact_articles:
            if len(article_compact) >= 4 and compact.startswith(article_compact):
                return article
        return None

    try:
        for page_number in range(len(document)):
            render_scales = (1.1, 0.9) if page_number == 1 else (1.1,)
            for render_scale in render_scales:
                page = document[page_number]
                try:
                    image = page.render(scale=render_scale).to_pil()
                finally:
                    if hasattr(page, "close"):
                        page.close()

                width, height = image.size
                crop_top = int(height * 0.39)
                table_image = image.crop((0, crop_top, width, int(height * 0.77)))
                words = []

                for box, text, confidence in iter_rapidocr_results(ocr, table_image):
                    if confidence < 0.30:
                        continue
                    raw_text = norm_text(text)
                    if not raw_text:
                        continue
                    xs = [point[0] for point in box]
                    ys = [point[1] + crop_top for point in box]
                    words.append(
                        {
                            "text": raw_text,
                            "x": min(xs),
                            "source_x": min(xs),
                            "y": -min(ys),
                        }
                    )

                for line in group_pdf_lines(words, y_tolerance=6):
                    line_words = sorted(line["words"], key=lambda item: item["x"])
                    text_line = normalize_maharaja_ocr_text(" ".join(word["text"] for word in line_words))
                    if not text_line or text_line.upper().startswith(("ORDER", "SUB", "TOTAL", "AMOUNT", "GOOD")):
                        continue

                    left_text = normalize_maharaja_ocr_text(" ".join(word["text"] for word in line_words if word["x"] < width * 0.60))
                    article = find_compact_article(left_text) or find_article_prefix(left_text, mapping)
                    if not article:
                        article = find_compact_article(text_line) or find_article_prefix(text_line, mapping)
                    if not article:
                        continue

                    qty_match = qty_re.search(text_line)
                    if qty_match:
                        quantity = qty_match.group("quantity")
                        after_qty = text_line[qty_match.end() :]
                    else:
                        quantity = ""
                        after_qty = text_line

                    numbers = amount_re.findall(after_qty)
                    if len(numbers) < 1:
                        continue
                    if "SUB TOTAL" in after_qty.upper() and len(numbers) >= 2:
                        unit_price, amount = numbers[0], numbers[1]
                    elif len(numbers) >= 2:
                        unit_price, amount = numbers[-2], numbers[-1]
                    else:
                        quantity_number = parse_pdf_number(quantity)
                        unit_price = ""
                        number = parse_pdf_number(numbers[0])
                        if isinstance(quantity_number, (int, float)) and quantity_number and isinstance(number, (int, float)):
                            if number <= 20:
                                unit_price = numbers[0]
                                amount = round(float(quantity_number) * float(number), 2)
                            else:
                                amount = numbers[0]
                                unit_price = f"{number / quantity_number:.2f}"
                        else:
                            amount = numbers[0]

                    if not quantity:
                        unit_price_number = parse_pdf_number(unit_price)
                        amount_number = parse_pdf_number(amount)
                        if (
                            isinstance(unit_price_number, (int, float))
                            and unit_price_number
                            and isinstance(amount_number, (int, float))
                        ):
                            derived_quantity = amount_number / unit_price_number
                            if abs(derived_quantity - round(derived_quantity)) < 0.01:
                                quantity = str(int(round(derived_quantity)))
                    if not quantity:
                        continue

                    article = canonical_article_key(article, mapping)
                    hs = mapping.get(article) or mapping.get(compact_key(article))
                    if not hs:
                        continue

                    key = (
                        page_number + 1,
                        article,
                        quantity,
                        str(parse_pdf_number(unit_price)),
                        str(parse_pdf_number(amount)),
                    )
                    if key in seen:
                        continue
                    seen.add(key)

                    rows.append(
                        {
                            "page": page_number + 1,
                            "line": int(-line["y"]),
                            "article": article,
                            "lookup_article": article,
                            "hs_code": hs,
                            "quantity": quantity,
                            "unit_price": unit_price,
                            "amount": amount,
                            "skip_value_fallback": True,
                            "raw_text": text_line,
                            "text": text_line,
                        }
                    )
    finally:
        document.close()

    return rows


def extract_pdf_articles_from_panache_exports(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"PANACHE\s+EXPORTS", full_text, flags=re.I):
            return []
        if not re.search(r"Order\s+No\.\s+Article\s+No\.\s+HSN\s+CODE", full_text, flags=re.I):
            return []
        if not re.search(r"QUANTITY\s+RATE\s+AMOUNT", full_text, flags=re.I):
            return []

        article_line_re = re.compile(
            r"^\s*(?P<order>\d{6})\s+(?P<article>[A-Z]?\d{3,6})\b\s+"
            r"(?:(?P<hsn>\d{7,10})\s+)?(?P<rest>.+)$",
            flags=re.I,
        )
        amount_re = re.compile(
            r"(?P<quantity>\d+)\s+(?P<unit_price>\d+(?:[,.]\d+)?)\s+"
            r"(?P<amount>\d[\d.]*[,.]\d{2})\s*$"
        )
        current = None

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            lines = group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            )

            for line in lines:
                text_line = norm_text(" ".join(word["text"] for word in line["words"]))
                if not text_line or re.search(r"TOTAL\s+PCS", text_line, flags=re.I):
                    continue

                article_match = article_line_re.match(text_line)
                if article_match:
                    article = norm_key(article_match.group("article"))
                    hs = mapping.get(article) or mapping.get(compact_key(article))
                    current = {
                        "page": page_number,
                        "line": int(-line["y"]),
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                    } if hs else None

                amount_match = amount_re.search(text_line)
                if not amount_match or not current:
                    continue

                row = dict(current)
                row["page"] = page_number
                row["line"] = int(-line["y"])
                row["quantity"] = amount_match.group("quantity")
                row["unit_price"] = amount_match.group("unit_price")
                row["amount"] = amount_match.group("amount")
                row["raw_text"] = text_line
                row["text"] = text_line
                rows.append(row)

    return rows


def extract_pdf_articles_from_silverline_rows(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join(get_pdf_page_text(page) for page in reader.pages)
    if not re.search(r"SILVERLINE\s+TACK\s+INC", full_text, flags=re.I):
        return []
    if not re.search(r"Description\s+of\s+Goods.*HSN.*Quantity.*Rate.*Amount", full_text, flags=re.I | re.S):
        return []

    rows = []
    line_pattern = re.compile(
        r"^\s*(?P<article>\d{3,5}\s+[A-Z]{1,6}\s+[A-Z0-9]+)\s+.+?\s+"
        r"(?P<hsn>\d{4})\s+(?P<net_weight>\d+(?:\.\d+)?)\s+"
        r"(?P<quantity>\d+)\s+(?:PCS|PAIR)\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+"
        r"(?P<amount>\d[\d,]*(?:\.\d{2})?)\s*$",
        flags=re.I,
    )

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            match = line_pattern.match(line)
            if not match:
                continue

            article = norm_key(match.group("article"))
            hs = mapping.get(article) or mapping.get(compact_key(article))
            if not hs:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "quantity": match.group("quantity"),
                    "unit_price": match.group("unit_price"),
                    "amount": match.group("amount"),
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_rafah_matrix(input_pdf: Path, mapping: Dict[str, str]) -> List[dict]:
    try:
        import pdfplumber
    except ImportError:
        return []

    rows = []

    with pdfplumber.open(str(input_pdf)) as pdf:
        full_text = "\n".join(page.extract_text(layout=True) or page.extract_text() or "" for page in pdf.pages)
        if not re.search(r"Rafah\s+International", full_text, flags=re.I):
            return []
        if not re.search(r"Description\s+Of\s+Goods.*Qty.*Rate.*Amount", full_text, flags=re.I | re.S):
            return []

        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(layout=True) or page.extract_text() or ""
            if is_packing_list_page(text):
                continue

            words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
            article_lines = []
            value_rows = []

            for line in group_pdf_lines(
                [{"text": word["text"], "x": float(word["x0"]), "source_x": float(word["x0"]), "y": -float(word["top"])} for word in words],
                y_tolerance=3,
            ):
                line_words = line["words"]
                text_line = norm_text(" ".join(word["text"] for word in line_words))
                article_match = re.match(r"^(?P<article>\d{3,5}\s+[A-Z]{1,6})\b", text_line)
                if article_match and re.search(r"\d+\s*/\s*\d+", text_line):
                    article = norm_key(article_match.group("article"))
                    size_qty_pairs = re.findall(r"\b\d+[A-Z]?\s*/\s*(\d+)\b", text_line)
                    quantity = sum(int(qty) for qty in size_qty_pairs)
                    article_lines.append({"article": article, "quantity": quantity, "text": text_line, "line": int(-line["y"])})

                right_numbers = [
                    word["text"]
                    for word in sorted(line_words, key=lambda item: item["x"])
                    if word["x"] > 430 and re.fullmatch(r"\d[\d,]*(?:\.\d+)?", word["text"])
                ]
                if len(right_numbers) >= 3:
                    qty, unit_price, amount = right_numbers[-3:]
                    value_rows.append({"quantity": int(qty.replace(",", "")), "unit_price": unit_price, "amount": amount})

            for article_line in article_lines:
                value_row = next((row for row in value_rows if row["quantity"] == article_line["quantity"]), None)
                if not value_row:
                    continue

                article = article_line["article"]
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if not hs:
                    continue

                rows.append(
                    {
                        "page": page_number,
                        "line": article_line["line"],
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "quantity": str(value_row["quantity"]),
                        "unit_price": value_row["unit_price"],
                        "amount": value_row["amount"],
                        "raw_text": article_line["text"],
                        "text": article_line["text"],
                    }
                )

    return rows


def extract_pdf_articles_from_item_code_rows(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join(get_pdf_page_text(page) for page in reader.pages)
    if not re.search(r"ITEM\s+DESCRIPTION\s+ITEM\s+CODE", full_text, flags=re.I):
        return []

    rows = []
    article_re = re.compile(r"\b(?P<article>\d{3,5}\s+[A-Z]{2,6}(?:\s+\d+)?)\b")

    def find_hs(article: str, line: str) -> Tuple[Optional[str], str]:
        article = norm_key(article)
        candidates = [article]
        mm_match = re.search(r"\b\d+\s*MM\b", line, flags=re.I)
        if mm_match:
            candidates.insert(0, norm_key(f"{article} {mm_match.group(0)}"))

        for candidate in candidates:
            hs = mapping.get(candidate) or mapping.get(compact_key(candidate))
            if hs:
                return hs, candidate

        compact_article = compact_key(article)
        prefixed = [
            key
            for key in mapping
            if len(key) > len(article) and compact_key(key).startswith(compact_article)
        ]
        if len(set(prefixed)) == 1:
            key = prefixed[0]
            return mapping[key], key
        return None, article

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            if not re.search(r"\bPCS\b", line, flags=re.I):
                continue
            article_match = article_re.search(line)
            qty_match = re.search(r"(?P<quantity>\d+)\s*PCS\b", line, flags=re.I)
            if not article_match or not qty_match:
                continue

            after_pcs = line[qty_match.end() :]
            numbers = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", after_pcs)
            if len(numbers) < 2:
                continue

            hs, lookup_article = find_hs(article_match.group("article"), line)
            if not hs:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": lookup_article,
                    "lookup_article": lookup_article,
                    "hs_code": hs,
                    "quantity": qty_match.group("quantity"),
                    "unit_price": numbers[0],
                    "amount": numbers[-1],
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_brading_grouped_totals(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    amount_pattern = re.compile(
        r"(?P<quantity>\d+)\s+\D*(?P<unit_price>\d+\.\d+)\s+\D*(?P<amount>\d[\d,]*\.\d{2})\s*$"
    )
    article_pattern_re = re.compile(r"^\s*(?P<order>\d{6})\s+(?P<article>\d{3,5})\b")

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue
        if page_number == 1 and not re.search(r"Order\s+Article\s+Item.*Total\s+Price\s+Amount", text, flags=re.I | re.S):
            continue

        current = None
        for line_number, line in enumerate(text.splitlines(), start=1):
            article_match = article_pattern_re.match(line)
            if article_match:
                article = norm_key(article_match.group("article"))
                hs = mapping.get(article) or mapping.get(compact_key(article))
                current = {
                    "page": page_number,
                    "line": line_number,
                    "article": article,
                    "lookup_article": article,
                    "hs_code": hs,
                    "raw_text": line,
                    "text": norm_text(line),
                } if hs else None

            amount_match = amount_pattern.search(line)
            if not amount_match or not current:
                continue

            row = dict(current)
            row["quantity"] = amount_match.group("quantity")
            row["unit_price"] = amount_match.group("unit_price")
            row["amount"] = amount_match.group("amount")
            row["raw_text"] = norm_text(f"{current.get('raw_text', '')} {line}")
            row["text"] = norm_text(f"{current.get('text', '')} {line}")
            rows.append(row)
            current = None

    return rows


def extract_pdf_articles_from_kartikeya_grouped_totals(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    full_text = "\n".join(get_pdf_page_text(page) for page in reader.pages)
    if not re.search(r"Kartikeya\s+International", full_text, flags=re.I):
        return []
    if not re.search(r"Our\s+Product\s+Code.*Quantity.*Rate.*Amount", full_text, flags=re.I | re.S):
        return []

    rows = []
    current = None
    article_pattern_re = re.compile(r"^\s*(?P<article>\d{3,5})\b")
    amount_pattern = re.compile(
        r"(?P<quantity>\d[\d,]*)\s*pc\s+(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>\d[\d,]*(?:\.\d{2})?)\s*$",
        flags=re.I,
    )

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        for line_number, line in enumerate(text.splitlines(), start=1):
            article_match = article_pattern_re.match(line)
            if article_match:
                article = norm_key(article_match.group("article"))
                hs = mapping.get(article) or mapping.get(compact_key(article))
                if hs:
                    current = {
                        "page": page_number,
                        "line": line_number,
                        "article": article,
                        "lookup_article": article,
                        "hs_code": hs,
                        "raw_text": line,
                        "text": norm_text(line),
                    }

            amount_match = amount_pattern.search(line)
            if not amount_match or not current:
                continue

            row = dict(current)
            row["quantity"] = amount_match.group("quantity").replace(",", "")
            row["unit_price"] = amount_match.group("unit_price")
            row["amount"] = amount_match.group("amount").replace(",", "")
            row["raw_text"] = norm_text(f"{current.get('raw_text', '')} {line}")
            row["text"] = norm_text(f"{current.get('text', '')} {line}")
            rows.append(row)
            current = None

    return rows


def find_product_code_bounds(header_line: str) -> Optional[Tuple[int, int]]:
    header_lower = header_line.lower()
    start = header_lower.find("our product code")
    if start < 0:
        return None

    next_headers = [
        header_lower.find("party's code"),
        header_lower.find("party code"),
        header_lower.find("order no"),
        header_lower.find("colour"),
        header_lower.find("color"),
    ]
    next_headers = [idx for idx in next_headers if idx > start]
    end = min(next_headers) if next_headers else start + 24
    return start, max(end, start + 12)


def extract_pdf_articles_left_of_order_no(
    reader: PdfReader,
    mapping: Dict[str, str],
    patterns: List[Tuple[str, re.Pattern]],
) -> List[dict]:
    rows = []

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        order_col: Optional[int] = None
        for line_number, line in enumerate(text.splitlines(), start=1):
            header_match = re.search(r"\bORD\.?\s*NO\.?", line, flags=re.I)
            if header_match and re.search(r"\bSIZE\b", line[: header_match.start()], flags=re.I):
                order_col = header_match.start()
                continue

            if order_col is None or not norm_text(line):
                continue

            normalized_line = norm_key(line)
            if normalized_line.startswith(("TOTAL", "AMOUNT", "DECLARATION", "THE EXPORTER", "NET WEIGHT", "GROSS WEIGHT")):
                continue

            left_of_order = line[:order_col].strip()
            article = find_article_in_text(left_of_order, mapping, patterns)
            if not article:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": display_article_code(left_of_order, article),
                    "lookup_article": article,
                    "hs_code": mapping[article],
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def find_party_code_bounds(header_line: str) -> Optional[Tuple[int, int]]:
    header_lower = header_line.lower()
    match = re.search(r"party['’]s\s+code", header_lower)
    if not match:
        return None

    start = max(0, match.start() - 2)
    next_headers = [
        header_lower.find("order no", match.end()),
        header_lower.find("order number", match.end()),
        header_lower.find("colour", match.end()),
        header_lower.find("color", match.end()),
    ]
    next_headers = [idx for idx in next_headers if idx > match.start()]
    end = min(next_headers) if next_headers else match.end() + 20
    return start, max(end, start + 12)


def find_buyer_product_code_bounds(header_line: str, next_line: str = "") -> Optional[Tuple[int, int]]:
    buyer_match = re.search(r"\bbuyer\b", header_line, flags=re.I)
    if not buyer_match:
        return None

    product_code_match = re.search(r"product\s+code", next_line, flags=re.I)
    if not product_code_match or abs(product_code_match.start() - buyer_match.start()) > 10:
        return None

    start = max(0, min(buyer_match.start(), product_code_match.start()) - 4)
    qty_match = re.search(r"\bqty\.?\b", header_line[buyer_match.end() :], flags=re.I)
    if qty_match:
        end = buyer_match.end() + qty_match.start()
    else:
        description_match = re.search(r"\bproduct\b", header_line[buyer_match.end() :], flags=re.I)
        end = buyer_match.end() + description_match.start() if description_match else start + 24
    return start, max(end, start + 14)


def extract_pdf_articles_from_buyer_product_code_column(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    last_bounds: Optional[Tuple[int, int]] = None

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        lines = text.splitlines()
        bounds = None
        header_index = None
        for idx, line in enumerate(lines):
            next_line = lines[idx + 1] if idx + 1 < len(lines) else ""
            bounds = find_buyer_product_code_bounds(line, next_line)
            if bounds:
                header_index = idx + 1
                last_bounds = bounds
                break

        if bounds is None or header_index is None:
            if last_bounds is None:
                continue
            bounds = last_bounds
            header_index = -1

        start, end = bounds
        for line_number, line in enumerate(lines[header_index + 1 :], start=header_index + 2):
            if not norm_text(line):
                continue

            normalized_line = norm_key(line)
            if normalized_line.startswith(("TOTAL", "GRAND TOTAL", "AMOUNT", "DECLARATION", "SIGNATURE", "NET WEIGHT", "GROSS WEIGHT")):
                continue

            cell = line[start:end].strip() if len(line) > start else ""
            article = find_article_in_text(cell, mapping, []) or find_article_prefix(cell, mapping)
            if not article:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": display_article_code(cell, article),
                    "lookup_article": article,
                    "hs_code": mapping[article],
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_party_code_column(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    last_bounds: Optional[Tuple[int, int]] = None

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        lines = text.splitlines()
        bounds = None
        header_index = None
        for idx, line in enumerate(lines):
            bounds = find_party_code_bounds(line)
            if bounds:
                header_index = idx
                last_bounds = bounds
                break

        if bounds is None or header_index is None:
            if last_bounds is None:
                continue
            bounds = last_bounds
            header_index = -1

        start, end = bounds
        for line_number, line in enumerate(lines[header_index + 1 :], start=header_index + 2):
            if not norm_text(line):
                continue

            normalized_line = norm_key(line)
            if normalized_line.startswith(("TOTAL", "GRAND TOTAL", "AMOUNT", "DECLARATION", "SIGNATURE", "GSTIN", "PAN")):
                continue

            cell = line[start:end].strip() if len(line) > start else ""
            article = find_article_in_text(cell, mapping, []) or find_article_prefix(cell, mapping)
            if not article:
                continue

            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": display_article_code(cell, article),
                    "lookup_article": article,
                    "hs_code": mapping[article],
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_product_code_column(reader: PdfReader, mapping: Dict[str, str]) -> List[dict]:
    rows = []
    seen = set()
    last_bounds: Optional[Tuple[int, int]] = None

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        lines = text.splitlines()
        bounds = None
        header_index = None
        for idx, line in enumerate(lines):
            bounds = find_product_code_bounds(line)
            if bounds:
                header_index = idx
                last_bounds = bounds
                break

        if bounds is None or header_index is None:
            if last_bounds is None:
                continue
            bounds = last_bounds
            header_index = -1

        start, end = bounds
        for line_number, line in enumerate(lines[header_index + 1 :], start=header_index + 2):
            if not norm_text(line):
                continue

            normalized_line = norm_key(line)
            if normalized_line.startswith(("TOTAL", "GRAND TOTAL", "AMOUNT", "DECLARATION", "SIGNATURE", "GSTIN", "PAN")):
                continue

            cell = line[start:end].strip() if len(line) > start else ""
            article = find_article_prefix(cell, mapping)
            if not article:
                continue

            key = (page_number, line_number, article)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "page": page_number,
                    "line": line_number,
                    "article": display_article_code(cell, article),
                    "lookup_article": article,
                    "hs_code": mapping[article],
                    "raw_text": line,
                    "text": norm_text(line),
                }
            )

    return rows


def extract_pdf_articles_from_positions(
    reader: PdfReader,
    mapping: Dict[str, str],
    patterns: List[Tuple[str, re.Pattern]],
) -> List[dict]:
    rows = []

    for page_number, page in enumerate(reader.pages, start=1):
        if is_packing_list_page(get_pdf_page_text(page)):
            continue
        lines = group_pdf_lines(extract_pdf_words(page))
        article_column = find_pdf_article_column(lines)
        if not article_column:
            continue

        for line_number, line in enumerate(lines, start=1):
            if line["y"] >= article_column["header_y"] - 2:
                continue

            column_words = [
                word["text"]
                for word in line["words"]
                if article_column["left"] <= word["source_x"] <= article_column["right"]
            ]
            column_text = norm_text(" ".join(column_words))
            if norm_key(column_text).startswith(("SUB TOTAL", "TOTAL", "EURO ", "N W", "NW", "G W", "GW")):
                continue

            matched_article = find_article_in_text(column_text, mapping, patterns)
            if matched_article:
                rows.append(
                    {
                        "page": page_number,
                        "line": line_number,
                        "article": display_article_code(column_text, matched_article),
                        "lookup_article": matched_article,
                        "hs_code": mapping[matched_article],
                        "text": line["text"],
                    }
                )

    return rows


def find_layout_article_bounds(header_line: str, next_line: str = "") -> Optional[Tuple[int, int]]:
    header_lower = header_line.lower()
    match = re.search(r"\b(?:article|art)\.?\s*(?:no\.?|number)\b", header_lower)
    if not match:
        article_match = re.search(r"\barticle\b", header_lower)
        if not article_match:
            return None

        next_lower = next_line.lower()
        number_matches = list(re.finditer(r"\bnumber\b", next_lower))
        aligned_number = None
        for number_match in number_matches:
            if abs(number_match.start() - article_match.start()) <= 8:
                aligned_number = number_match
                break
        if not aligned_number:
            return None

        match = article_match

    short_art_header = bool(re.fullmatch(r"art\.?\s*no\.?", match.group(0), flags=re.I))
    start = max(0, match.start() - (10 if short_art_header else 3))
    next_matches = [
        m.start()
        for m in re.finditer(
            r"\b(item|qty|quantity|hsn|hs\s*code|price|amount|total|cartons?|ctn|pcs|description|desc)\b",
            header_lower[match.end() :],
        )
    ]
    end = match.end() + min(next_matches) if next_matches else len(header_line)
    return start, max(end, start + 18)


def extract_pdf_articles_from_layout(
    reader: PdfReader,
    mapping: Dict[str, str],
    patterns: List[Tuple[str, re.Pattern]],
) -> List[dict]:
    rows = []
    last_bounds: Optional[Tuple[int, int]] = None

    for page_number, page in enumerate(reader.pages, start=1):
        text = get_pdf_page_text(page)
        if is_packing_list_page(text):
            continue

        lines = text.splitlines()
        bounds = None
        header_index = None
        for idx, line in enumerate(lines):
            next_line = lines[idx + 1] if idx + 1 < len(lines) else ""
            bounds = find_layout_article_bounds(line, next_line)
            if bounds:
                header_index = idx
                last_bounds = bounds
                break

        if bounds is None or header_index is None:
            if last_bounds is None:
                continue
            bounds = last_bounds
            header_index = -1

        start, end = bounds
        for line_number, line in enumerate(lines[header_index + 1 :], start=header_index + 2):
            if not norm_text(line):
                continue

            normalized_line = norm_key(line)
            if normalized_line.startswith(("SUB TOTAL", "TOTAL", "EURO ", "N W", "NW", "G W", "GW")):
                continue

            article_cell = line[start:end].strip() if len(line) > start else ""
            matched_article = find_article_in_text(article_cell, mapping, patterns)
            if matched_article:
                rows.append(
                    {
                        "page": page_number,
                        "line": line_number,
                        "article": display_article_code(article_cell, matched_article),
                        "lookup_article": matched_article,
                        "hs_code": mapping[matched_article],
                        "raw_text": line,
                        "text": norm_text(line),
                    }
                )

    return rows


def extract_pdf_articles_from_ocr(
    input_pdf: Path,
    mapping: Dict[str, str],
) -> List[dict]:
    try:
        import pypdfium2 as pdfium
        ocr = create_rapidocr_engine()
    except ImportError as exc:
        raise RuntimeError(
            f"OCR kon niet worden gestart. Werkelijke fout: {exc}"
        ) from exc

    rows = []
    seen = set()
    text_reader = PdfReader(BytesIO(input_pdf.read_bytes()))
    document = pdfium.PdfDocument(str(input_pdf))

    try:
        for page_number in range(len(document)):
            if page_number < len(text_reader.pages) and is_packing_list_page(get_pdf_page_text(text_reader.pages[page_number])):
                continue

            page = document[page_number]
            try:
                image = page.render(scale=2).to_pil()
            finally:
                if hasattr(page, "close"):
                    page.close()

            width, height = image.size
            table_image = image.crop(
                (
                    max(0, int(width * 0.08)),
                    int(height * 0.40),
                    min(width, int(width * 0.96)),
                    min(height, int(height * 0.78)),
                )
            )
            for box, text, confidence in iter_rapidocr_results(ocr, table_image):
                if confidence < 0.75:
                    continue

                xs = [point[0] for point in box]
                ys = [point[1] for point in box]
                x_min = min(xs)
                y_min = min(ys)
                raw_text = norm_text(text)
                normalized = norm_key(raw_text)

                if not raw_text:
                    continue
                if x_min < 20 or x_min > 650:
                    continue
                if normalized.startswith(("ORDER", "SUBTOTAL", "TOTAL", "GROSS", "AMOUNT", "MARKS", "DESCRIPTION")):
                    continue

                article = find_article_prefix(raw_text, mapping)
                if not article:
                    continue

                key = (page_number + 1, round(y_min / 8), article)
                if key in seen:
                    continue
                seen.add(key)
                rows.append(
                    {
                        "page": page_number + 1,
                        "line": int(y_min),
                        "article": display_article_code(raw_text, article),
                        "lookup_article": article,
                        "hs_code": mapping[article],
                        "text": raw_text,
                    }
                )
    finally:
        document.close()

    return rows


def extract_pdf_articles(input_pdf: Path, mapping_csv: Path) -> List[dict]:
    mapping, names = load_catalog(mapping_csv)
    patterns = [(article, article_pattern(article)) for article in sorted(mapping, key=len, reverse=True)]
    reader = PdfReader(BytesIO(input_pdf.read_bytes()))

    rows = extract_pdf_articles_from_mehra_shoes(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_leather_art_variants(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_gng_pet_rows(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_zhongshan_biaoqi(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_changzhou_biseen(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_quanzhou_xingye(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_xiamen_grassland(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_changzhou_ziyuan_rows(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_karan_letex_rows(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_jone_shou_scan(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_maharaja_scan(input_pdf, reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_mark_equestrian_party_code(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_ibrahim_buyer_code(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_ruksh_rows(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_tarun_thermoware(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_panache_exports(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_shipment_columns(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_silverline_rows(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_rafah_matrix(input_pdf, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_item_code_rows(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_brading_grouped_totals(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_kartikeya_grouped_totals(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    buyer_product_rows = extract_pdf_articles_from_buyer_product_code_column(reader, mapping)
    party_rows = extract_pdf_articles_from_party_code_column(reader, mapping)
    product_rows = extract_pdf_articles_from_product_code_column(reader, mapping)
    preferred_rows = max((buyer_product_rows, party_rows, product_rows), key=len)
    if preferred_rows:
        return enrich_pdf_rows(preferred_rows, names)
    rows = extract_pdf_articles_left_of_order_no(reader, mapping, patterns)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_po_lines(reader, mapping)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_positions(reader, mapping, patterns)
    if rows:
        return enrich_pdf_rows(rows, names)
    rows = extract_pdf_articles_from_layout(reader, mapping, patterns)
    if rows:
        return enrich_pdf_rows(rows, names)
    return enrich_pdf_rows(extract_pdf_articles_from_ocr(input_pdf, mapping), names)


def fill_pdf_invoice(input_pdf: Path, output_xlsx: Path, mapping_csv: Path) -> dict:
    rows = extract_pdf_articles(input_pdf, mapping_csv)
    wb = Workbook()
    ws = wb.active
    ws.title = "HS codes"
    ws.append(
        [
            "Page",
            "Article No.",
            "Article name",
            "HS code",
            "Aantallen",
            "Waarde",
            "Totaal waarde",
            "PDF text",
        ]
    )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    group_fills = [
        PatternFill("solid", fgColor="EAF3F8"),
        PatternFill("solid", fgColor="FCE4D6"),
        PatternFill("solid", fgColor="E2F0D9"),
        PatternFill("solid", fgColor="FFF2CC"),
        PatternFill("solid", fgColor="EADCF8"),
        PatternFill("solid", fgColor="DDEBF7"),
    ]
    group_fill_by_key: Dict[str, PatternFill] = {}

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append(
            [
                row["page"],
                row["article"],
                row.get("article_name", ""),
                row["hs_code"],
                parse_pdf_number(row.get("quantity", "")),
                parse_pdf_number(row.get("unit_price", "")),
                parse_pdf_number(row.get("amount", "")),
                row["text"],
            ]
        )
        excel_row = ws.max_row
        group_key = article_group_key(row.get("article", ""))
        if group_key not in group_fill_by_key:
            group_fill_by_key[group_key] = group_fills[len(group_fill_by_key) % len(group_fills)]
        row_fill = group_fill_by_key[group_key]

        for col_idx in range(1, 9):
            ws.cell(excel_row, col_idx).fill = row_fill

        if row.get("quantity") or row.get("unit_price") or row.get("amount"):
            for col_idx in range(5, 8):
                ws.cell(excel_row, col_idx).font = Font(bold=True)
            ws.cell(excel_row, 5).number_format = "#,##0"
            ws.cell(excel_row, 6).number_format = "#,##0.00"
            ws.cell(excel_row, 7).number_format = "#,##0.00"

    if rows:
        total_row = ws.max_row + 1
        invoice_total = sum(
            float(ws.cell(row_idx, 7).value or 0)
            for row_idx in range(2, total_row)
            if isinstance(ws.cell(row_idx, 7).value, (int, float))
        )
        ws.cell(total_row, 6).value = "Factuur waarde"
        ws.cell(total_row, 7).value = invoice_total
        total_fill = PatternFill("solid", fgColor="D9EAD3")
        for col_idx in range(1, 9):
            cell = ws.cell(total_row, col_idx)
            cell.fill = total_fill
            cell.font = Font(bold=True)
        ws.cell(total_row, 7).number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 100
    wb.save(output_xlsx)
    return {"filled": len(rows), "unmatched": [], "unmatched_count": 0}


if __name__ == "__main__":
    if len(sys.argv) == 1:
        runpy.run_path(str(Path(__file__).with_name("app.py")), run_name="__main__")
        raise SystemExit(0)

    if len(sys.argv) not in (3, 4):
        print("Gebruik: python fill_hs_codes.py invoice.xlsx output.xlsx [hs_mapping.csv]")
        raise SystemExit(2)
    input_file = Path(sys.argv[1])
    output_file = Path(sys.argv[2])
    mapping_file = Path(sys.argv[3]) if len(sys.argv) == 4 else Path(__file__).with_name("hs_mapping.csv")
    result = fill_invoice(input_file, output_file, mapping_file)
    print(f"Klaar: {result['filled']} HS-codes ingevuld. Output: {output_file}")
    if result["unmatched_count"]:
        print(f"Niet gevonden: {result['unmatched_count']} artikelen. Eerste voorbeelden:")
        for item in result["unmatched"][:20]:
            print(f"- {item['sheet']} rij {item['row']}: {item['article']}")
